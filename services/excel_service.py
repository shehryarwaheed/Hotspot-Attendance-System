import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def parse_student_excel(file_content: bytes):
    df = pd.read_excel(BytesIO(file_content))
    students = []
    
    # Normalize columns to lowercase for easier matching
    orig_cols = list(df.columns)
    low_cols = [str(c).lower().strip() for c in orig_cols]
    
    # Try to find Roll Number and Name columns
    roll_idx = next((i for i, c in enumerate(low_cols) if 'roll' in c or 'id' in c), 0)
    name_idx = next((i for i, c in enumerate(low_cols) if 'name' in c or 'student' in c), 1)

    for _, row in df.iterrows():
        try:
            roll_val = str(row.iloc[roll_idx]).strip()
            name_val = str(row.iloc[name_idx]).strip()
            
            if roll_val and name_val and roll_val.lower() != "nan" and name_val.lower() != "nan":
                students.append({"roll_number": roll_val, "name": name_val})
        except:
            continue
    return students

def export_attendance_to_excel(attendance_data, section, date, lecture_title):
    """
    Standard single-session or flat export. 
    Now upgraded to use openpyxl for better styling (Green P / Red A).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header Styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Determine columns based on first data row
    if not attendance_data:
        headers = ["Roll Number", "Name", "Section", "Status"]
    else:
        headers = list(attendance_data[0].keys())

    # Write Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Write Data
    p_total = 0
    a_total = 0
    for row_idx, row_data in enumerate(attendance_data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            
            # Apply color coding if it's a Status or Date column
            if val == 'P' or val == 'present':
                cell.value = 'P'
                cell.font = Font(color="059669", bold=True) # Green
                if header.lower() == "status": p_total += 1
            elif val == 'A' or val == 'absent':
                cell.value = 'A'
                cell.font = Font(color="DC2626", bold=True) # Red
                if header.lower() == "status": a_total += 1
            
            cell.alignment = center_align

    # 4. Summary Totals at Bottom
    last_row = len(attendance_data) + 3
    status_col_idx = next((i for i, h in enumerate(headers, 1) if h.lower() == "status"), 4)
    
    # Total Present Row
    ws.cell(row=last_row, column=status_col_idx - 1, value="TOTAL PRESENT").font = Font(bold=True)
    cp = ws.cell(row=last_row, column=status_col_idx, value=p_total)
    cp.font = Font(color="059669", bold=True)
    cp.alignment = center_align

    # Total Absent Row
    ws.cell(row=last_row + 1, column=status_col_idx - 1, value="TOTAL ABSENT").font = Font(bold=True)
    ca = ws.cell(row=last_row + 1, column=status_col_idx, value=a_total)
    ca.font = Font(color="DC2626", bold=True)
    ca.alignment = center_align

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Filename logic
    if date:
        filename = f"Attendance_{section}_{date}_{lecture_title}.xlsx"
    else:
        filename = f"Full_Semester_Attendance_{section}_{lecture_title}.xlsx"
    
    for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        filename = filename.replace(char, "_")

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename

def export_matrix_attendance_to_excel(matrix_data, section, lecture_title, dates):
    """
    Professional Matrix Export: 
    Roll | Name | Section | Date1 | Date2 | ... | Total P | Total A | %
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Semester Report"

    # 1. Header Section
    ws.merge_cells('A1:E1')
    ws['A1'] = "ATTENDANCE REPORT"
    ws['A1'].font = Font(size=16, bold=True, color="4F46E5")
    ws['A1'].alignment = Alignment(horizontal="left")

    ws['A2'] = f"Subject: {lecture_title}"
    ws['A2'].font = Font(bold=True)
    ws['A3'] = f"Section: {section}"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = f"Total Lectures: {len(dates)}"
    ws['A4'].font = Font(bold=True)

    # 2. Table Headers
    start_row = 6
    headers = ["Roll Number", "Name", "Section"] + dates + ["Total P", "Total A"]
    
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 3. Data Rows
    for r_idx, row_data in enumerate(matrix_data, start_row + 1):
        p_count = 0
        a_count = 0
        
        # Static Columns
        ws.cell(row=r_idx, column=1, value=row_data["Roll Number"]).border = thin_border
        ws.cell(row=r_idx, column=2, value=row_data["Name"]).border = thin_border
        ws.cell(row=r_idx, column=3, value=row_data["Section"]).border = thin_border

        # Date Columns
        for c_idx, date_str in enumerate(dates, 4):
            status = row_data.get(date_str, "A")
            cell = ws.cell(row=r_idx, column=c_idx, value=status)
            cell.alignment = center_align
            cell.border = thin_border
            
            if status == 'P':
                cell.font = Font(color="059669", bold=True) # Green
                p_count += 1
            else:
                cell.font = Font(color="DC2626", bold=True) # Red
                a_count += 1
        
        # Summary Columns
        total_col = 4 + len(dates)
        cell_p = ws.cell(row=r_idx, column=total_col, value=p_count)
        cell_p.border = thin_border
        cell_p.alignment = center_align

        cell_a = ws.cell(row=r_idx, column=total_col + 1, value=a_count)
        cell_a.border = thin_border
        cell_a.alignment = center_align

    # 4. Daily Column Totals at Bottom
    summary_row_start = start_row + len(matrix_data) + 2
    
    ws.cell(row=summary_row_start, column=3, value="TOTAL PRESENT").font = Font(bold=True)
    ws.cell(row=summary_row_start + 1, column=3, value="TOTAL ABSENT").font = Font(bold=True)

    for c_idx, date_str in enumerate(dates, 4):
        daily_p = sum(1 for row in matrix_data if row.get(date_str) == 'P')
        daily_a = sum(1 for row in matrix_data if row.get(date_str) == 'A')
        
        cp = ws.cell(row=summary_row_start, column=c_idx, value=daily_p)
        cp.font = Font(color="059669", bold=True)
        cp.alignment = center_align
        cp.border = thin_border

        ca = ws.cell(row=summary_row_start + 1, column=c_idx, value=daily_a)
        ca.font = Font(color="DC2626", bold=True)
        ca.alignment = center_align
        ca.border = thin_border

    # Adjust width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40 # Expanded Name column
    ws.column_dimensions['C'].width = 12

    filename = f"Report_{section}_{lecture_title}.xlsx"
    for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        filename = filename.replace(char, "_")

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename
